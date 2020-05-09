import re
from copy import copy
import pandas as pd
# from openpyxl.reader.excel import load_workbook
# from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows




# def copyRange(startCol, startRow, endCol, endRow, sheet):
#     rangeSelected = []
#     for i in range(startRow, endRow + 1, 1):
#         rowSelected = []
#         for j in range(startCol, endCol + 1, 1):
#             rowSelected.append(sheet.cell(row=i, column=j).value)
#         rangeSelected.append(rowSelected)
#
#     return rangeSelected
#
#
# def pasteRange(startCol, startRow, sheetReceiving, copiedData):
#
#     for i, row in enumerate(copiedData):
#         for j, col in enumerate(row):
#             sheetReceiving.cell(row=startRow+i, column=startCol+j).value = col
#     return


def spryreport(template_filename, data):
    """
    Отчет из шаблона. Переменные в {{переменная}}

    :param template_filename: Шаблон
    :param data: Первый уровень лист, второй уровень переменные на листе({} шапка, [] спецификация).
    :return: Возвращает открытую книгу book, которую можно сохранить или отправить в поток.
    """
    if not isinstance(data, dict):
        return Exception("Data must be dict")

    book = load_workbook(template_filename)
    sheets = book.sheetnames  # Список листов
    # print(sheets)
    for sheet_name in data:  # Верхний уровень входных данных - наименование листа в excel
        # print (sheet_name)
        sheet_excel = None
        for i, v in enumerate(sheets):
            if v.lower() == sheet_name.lower():
                sheet_excel = book[sheets[i]]  # Берем по имени лист



        if sheet_excel is not None:
            dict_for_this_sheet = data[sheet_name]  # Смотрим данные только для этого листа

            df = pd.DataFrame(sheet_excel.values)  # Считываем лист xlsx
            # print(df)

            patterns = []
            v_datas = []
            for k_data, v_data in dict_for_this_sheet.items():
                if not (isinstance(v_data, (list, dict)) or v_data is None):  # Не листы и не пустое значение
                    pattern = re.compile(re.escape('{{' + k_data + '}}'), re.IGNORECASE)
                    patterns.append(pattern)
                    v_datas.append(str(v_data))
                    # df.replace(pattern, str(v_data), inplace=True, regex=True)  # Делам в DataFrame замену по патерну то что пришло из data дл этого листа

            df.replace(patterns, v_datas, inplace=True, regex=True)  # Делам в DataFrame замену по патерну то что пришло из data дл этого листа
            rows = dataframe_to_rows(df, index=False, header=False)  # Проставляем назад в эксель.

            for c_row, row in enumerate(rows):
                for c_col, col in enumerate(row):
                    sheet_excel.cell(row=c_row + 1, column=c_col + 1, value=col)


            #Код для листов. Ищем их сопостовляем колонки
            items_meta = []
            for k_data, v_data in dict_for_this_sheet.items(): #Ищем детализацию
                if isinstance(v_data, list):
                    for i_row, v_row in enumerate(sheet_excel.rows): #Ищем строку где детализация
                        for i_cell, cell in enumerate(v_row): #Перебираем все ячейки в строке
                            if isinstance(cell.value, str):
                                # if cell.value.lower().find('{{'+k_data.lower()+'\..*}}')>=0:
                                item = re.search('{{' + k_data.lower() + '\..*}}', cell.value.lower())
                                if item:
                                    item_last = re.search('\..*}}', item.group(0)).group(0)[1:-2]
                                    # print(i_cell, item.group(0), item_last)
                                    if cell.value.lower() == item.group(0).lower():
                                        pattern = None
                                    else :
                                        pattern = re.compile(re.escape(item.group(0)), re.IGNORECASE)
                                    items_meta.append({'cell':cell,
                                                       'replaced': pattern,
                                                       'field':item_last,
                                                       'template':cell.value})


                    if items_meta:
                        append_row(items_meta, v_data, sheet_excel)
                    break

    return book


def append_row(items_meta, data, sheet_excel):
    """

    :param items_meta: Сопостовление полей
    :param data: Список объектов поле и значения
    :param sheet_excel: Открытый лист
    :return:
    """
    first_row = items_meta[0]['cell'].row
    sheet_excel.insert_rows(first_row + 1, len(data))


    for i_data, v_data in enumerate(data):
        for i_meta, v_meta in enumerate(items_meta):
            value = None
            for k, v in v_data.items():
                if k.lower() == v_meta['field']:
                    if v_meta['replaced']:
                        value = v_meta['replaced'].sub(str(v), v_meta['template'])
                    else:
                        value = v

                    break
            new_cell = sheet_excel.cell(row=first_row+i_data, column=v_meta['cell'].col_idx, value=(value))
            if v_meta['cell'].has_style:
                new_cell.font = copy(v_meta['cell'].font)
                new_cell.border = copy(v_meta['cell'].border)
                new_cell.fill = copy(v_meta['cell'].fill)
                new_cell.number_format = copy(v_meta['cell'].number_format)
                new_cell.protection = copy(v_meta['cell'].protection)
                new_cell.alignment = copy(v_meta['cell'].alignment)


    return